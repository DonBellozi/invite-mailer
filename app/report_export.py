from __future__ import annotations

import html
import io
import re
from dataclasses import dataclass
from datetime import datetime
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle
from ttf_opensans import opensans

from .audience import is_explicit_template
from .db import Database
from .indigo import summarize_employee_result
from .placements import eligible_placements


FONT_REGULAR = "ReportOpenSans"
FONT_BOLD = "ReportOpenSansBold"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MEDIA_TYPE = "application/pdf"


@dataclass(frozen=True)
class ExportRow:
    fio: str
    email: str
    login: str
    department: str
    position: str
    status: str
    grade: str
    status_key: str


@dataclass(frozen=True)
class TestExportReport:
    template_id: str
    test_name: str
    created_at: datetime
    participants: int
    completed: int
    failed: int
    waiting: int
    rows: tuple[ExportRow, ...]


def _fmt_date(value: str | None, include_time: bool = True) -> str:
    if not value:
        return ""
    try:
        pattern = "%d.%m.%Y %H:%M" if include_time else "%d.%m.%Y"
        return datetime.fromisoformat(value).strftime(pattern)
    except ValueError:
        return str(value)


def _report_employee_entries(connection, template: dict):
    if is_explicit_template(template):
        employees = connection.execute(
            """
            SELECT e.*
            FROM test_assignments a
            JOIN employees e
              ON e.worker_key = a.worker_key
             AND e.employment_seq = a.employment_seq
            WHERE a.template_id = ?
              AND a.active = 1
              AND e.active = 1
            ORDER BY COALESCE(e.fio, '')
            """,
            (template["id"],),
        ).fetchall()
    else:
        employees = connection.execute(
            """SELECT * FROM employees WHERE active = 1
               ORDER BY COALESCE(fio, '')"""
        ).fetchall()

    entries = []
    for employee in employees:
        placements = eligible_placements(connection, employee, template)
        if placements:
            entries.append((employee, placements))
    entries.sort(key=lambda item: ((item[1][0].department if item[1] else "").casefold(), str(item[0]["fio"] or "").casefold()))
    return entries


def build_test_export_report(db: Database, template: dict, created_at: datetime | None = None) -> TestExportReport:
    export_time = created_at or datetime.now()
    completed_count = 0
    failed_count = 0
    participant_count = 0
    export_rows: list[ExportRow] = []

    with db.connect() as connection:
        entries = _report_employee_entries(connection, template)
        for employee, placements in entries:
            participant_count += 1
            result = summarize_employee_result(connection, employee, template)
            if result.status == "completed":
                completed_count += 1
                completion_date = _fmt_date(result.completed_at, include_time=False)
                status_text = f"Пройден ({completion_date})" if completion_date else "Пройден"
                grade = result.grade or ""
                status_key = "completed"
            elif result.status == "failed":
                failed_count += 1
                status_text = "Не прошел"
                grade = "Не прошел"
                status_key = "failed"
            else:
                status_text = "Ожидает прохождения"
                grade = ""
                status_key = "waiting"

            for placement in placements:
                export_rows.append(
                    ExportRow(
                        fio=str(employee["fio"] or ""),
                        email=str(employee["email"] or ""),
                        login=str(employee["login"] or ""),
                        department=placement.department,
                        position=placement.position,
                        status=status_text,
                        grade=grade,
                        status_key=status_key,
                    )
                )

    waiting_count = max(0, participant_count - completed_count - failed_count)
    return TestExportReport(
        template_id=str(template["id"]),
        test_name=str(template.get("name", template["id"])),
        created_at=export_time,
        participants=participant_count,
        completed=completed_count,
        failed=failed_count,
        waiting=waiting_count,
        rows=tuple(export_rows),
    )


def _safe_test_name(value: str) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", value)
    name = re.sub(r"\s+", "_", name)
    name = re.sub(r"_+", "_", name)
    return name.strip(" ._") or "Отчет"


def export_filename(report: TestExportReport, extension: str) -> str:
    safe_name = _safe_test_name(report.test_name)
    timestamp = report.created_at.strftime("%Y%m%d%H%M")
    return f"{safe_name}_{timestamp}.{extension.lstrip('.')}"


def download_headers(filename: str) -> dict[str, str]:
    fallback = re.sub(r"[^A-Za-z0-9._-]+", "_", filename)
    encoded = quote(filename, safe="")
    return {
        "Content-Disposition": f'attachment; filename="{fallback}"; filename*=UTF-8\'\'{encoded}',
        "Cache-Control": "no-store",
    }


def _metadata_lines(report: TestExportReport) -> tuple[str, str, str, str]:
    return (
        "Отчет о прохождении теста",
        f"Тест: {report.test_name}",
        f"Дата составления отчета: {report.created_at:%d.%m.%Y %H:%M}",
        f"Участников: {report.participants} | Пройдено: {report.completed} | Не прошли: {report.failed} | Ожидают прохождения: {report.waiting}",
    )


def build_xlsx(report: TestExportReport) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    sheet_name = re.sub(r"[\[\]:*?/\\]+", "_", report.test_name).strip()
    worksheet.title = sheet_name[:31] if sheet_name else "Отчет"
    worksheet.sheet_view.showGridLines = False

    title_text, test_text, date_text, stats_text = _metadata_lines(report)
    for row_number in range(1, 5):
        worksheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=7)
    worksheet["A1"], worksheet["A2"], worksheet["A3"], worksheet["A4"] = title_text, test_text, date_text, stats_text
    worksheet["A1"].font = Font(name="Arial", size=16, bold=True, color="1F2937")
    worksheet["A2"].font = Font(name="Arial", size=13, bold=True, color="1F2937")
    worksheet["A3"].font = Font(name="Arial", size=10, color="475467")
    worksheet["A4"].font = Font(name="Arial", size=10, bold=True, color="344054")
    for cell_name in ("A1", "A2", "A3", "A4"):
        worksheet[cell_name].alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 26
    worksheet.row_dimensions[2].height = 22
    worksheet.row_dimensions[3].height = 20
    worksheet.row_dimensions[4].height = 22

    header_row = 6
    headers = ("ФИО", "E-mail", "Логин", "Подразделение", "Должность", "Статус", "Оценка")
    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    header_font = Font(name="Arial", size=10, bold=True, color="1F2937")
    thin_side = Side(style="thin", color="D0D5DD")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for column, value in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column, value=value)
        cell.fill, cell.font, cell.border = header_fill, header_font, cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[header_row].height = 28

    status_colors = {"completed": "176B36", "failed": "B42318", "waiting": "8A5B00"}
    first_data_row = header_row + 1
    for row_offset, item in enumerate(report.rows):
        row_number = first_data_row + row_offset
        values = (item.fio, item.email, item.login, item.department, item.position, item.status, item.grade)
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_number, column=column, value=value)
            cell.font = Font(name="Arial", size=9, color="222222")
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        worksheet.cell(row=row_number, column=6).font = Font(name="Arial", size=9, bold=True, color=status_colors[item.status_key])

    for column, width in {"A":35,"B":30,"C":20,"D":50,"E":42,"F":26,"G":18}.items():
        worksheet.column_dimensions[column].width = width
    last_row = max(header_row, worksheet.max_row)
    worksheet.freeze_panes = f"A{first_data_row}"
    worksheet.auto_filter.ref = f"A{header_row}:G{last_row}"
    worksheet.print_title_rows = f"1:{header_row}"
    worksheet.print_area = f"A1:G{last_row}"
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.4
    worksheet.page_margins.bottom = 0.4
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _register_pdf_fonts() -> None:
    registered = set(pdfmetrics.getRegisteredFontNames())
    if FONT_REGULAR not in registered:
        regular_font = opensans(font_weight=400)
        pdfmetrics.registerFont(TTFont(FONT_REGULAR, str(regular_font.path)))
    if FONT_BOLD not in registered:
        bold_font = opensans(font_weight=700)
        pdfmetrics.registerFont(TTFont(FONT_BOLD, str(bold_font.path)))


def _pdf_paragraph(value: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(html.escape(value or "").replace("\n", "<br/>"), style)


def build_pdf(report: TestExportReport) -> bytes:
    _register_pdf_fonts()
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm,
        topMargin=10*mm, bottomMargin=12*mm,
        title=f"Отчет о прохождении теста {report.test_name}", author="Система тестирования",
    )
    title_style = ParagraphStyle(name="ReportTitle", fontName=FONT_BOLD, fontSize=16, leading=20, alignment=TA_LEFT, textColor=colors.HexColor("#1F2937"), spaceAfter=4*mm)
    test_style = ParagraphStyle(name="ReportTest", fontName=FONT_BOLD, fontSize=12, leading=15, alignment=TA_LEFT, textColor=colors.HexColor("#1F2937"), spaceAfter=2*mm)
    metadata_style = ParagraphStyle(name="ReportMetadata", fontName=FONT_REGULAR, fontSize=9, leading=12, alignment=TA_LEFT, textColor=colors.HexColor("#475467"), spaceAfter=1.5*mm)
    stats_style = ParagraphStyle(name="ReportStats", fontName=FONT_BOLD, fontSize=9, leading=12, alignment=TA_LEFT, textColor=colors.HexColor("#344054"), spaceAfter=3*mm)
    table_header_style = ParagraphStyle(name="TableHeader", fontName=FONT_BOLD, fontSize=7.5, leading=9, alignment=TA_CENTER, textColor=colors.HexColor("#1F2937"))
    table_cell_style = ParagraphStyle(name="TableCell", fontName=FONT_REGULAR, fontSize=7.4, leading=9.2, alignment=TA_LEFT, textColor=colors.HexColor("#222222"), splitLongWords=True)
    completed_style = ParagraphStyle(name="TableCompleted", parent=table_cell_style, fontName=FONT_BOLD, textColor=colors.HexColor("#176B36"))
    failed_style = ParagraphStyle(name="TableFailed", parent=table_cell_style, fontName=FONT_BOLD, textColor=colors.HexColor("#B42318"))
    waiting_style = ParagraphStyle(name="TableWaiting", parent=table_cell_style, fontName=FONT_BOLD, textColor=colors.HexColor("#8A5B00"))
    status_styles = {"completed":completed_style,"failed":failed_style,"waiting":waiting_style}

    title_text, test_text, date_text, stats_text = _metadata_lines(report)
    story = [
        _pdf_paragraph(title_text,title_style), _pdf_paragraph(test_text,test_style),
        _pdf_paragraph(date_text,metadata_style), _pdf_paragraph(stats_text,stats_style), Spacer(1,2*mm),
    ]
    table_data = [[_pdf_paragraph(value, table_header_style) for value in ("ФИО","E-mail","Логин","Подразделение","Должность","Статус","Оценка")]]
    for item in report.rows:
        table_data.append([
            _pdf_paragraph(item.fio,table_cell_style), _pdf_paragraph(item.email,table_cell_style),
            _pdf_paragraph(item.login,table_cell_style), _pdf_paragraph(item.department,table_cell_style),
            _pdf_paragraph(item.position,table_cell_style), _pdf_paragraph(item.status,status_styles[item.status_key]),
            _pdf_paragraph(item.grade,table_cell_style),
        ])
    table = LongTable(table_data, colWidths=[40*mm,42*mm,24*mm,48*mm,55*mm,38*mm,22*mm], repeatRows=1, hAlign="LEFT", splitByRow=1)
    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#DCE6F1")),
        ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#BFC7D1")),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",(0,0),(-1,-1),4), ("RIGHTPADDING",(0,0),(-1,-1),4),
        ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#FAFBFC")]),
    ]))
    story.append(table)

    def draw_footer(canvas, doc) -> None:
        canvas.saveState(); canvas.setFont(FONT_REGULAR,7); canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawRightString(landscape(A4)[0]-10*mm,6*mm,f"Страница {doc.page}"); canvas.restoreState()

    document.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    return output.getvalue()
