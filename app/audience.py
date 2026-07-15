from __future__ import annotations

import hashlib
import io
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook

from .db import Database
from .settings import Settings
from .xlsx_parser import normalize_email


EMAIL_HEADER_VARIANTS = {
    "email": 100,
    "e-mail": 100,
    "e mail": 95,
    "адрес электронной почты": 100,
    "электронная почта": 95,
    "электронный адрес": 90,
    "почта": 60,
}
SAFE_FILENAME_RE = re.compile(r"[^0-9A-Za-zА-Яа-я._-]+")

STATUS_LABELS = {
    "ready": "Готово к назначению",
    "imported": "Назначение создано",
    "already_assigned": "Уже назначен ранее",
    "duplicate": "Дубликат в файле",
    "invalid": "Некорректный email",
    "not_found": "Не найден в основной базе",
    "multiple_matches": "Несколько совпадений",
    "inactive_match": "Найден только неактивный сотрудник",
    "no_login": "Нет логина",
    "excluded_by_operator": "Исключено оператором",
}


class AudienceImportError(ValueError):
    pass


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def is_explicit_template(template: dict[str, Any]) -> bool:
    audience = template.get("audience") or {}
    return str(audience.get("type", "")).strip().lower() == "explicit_list"


def explicit_templates(settings: Settings) -> list[dict[str, Any]]:
    return [template for template in settings.templates if is_explicit_template(template)]


def get_template(settings: Settings, template_id: str) -> dict[str, Any]:
    for template in settings.templates:
        if template.get("id") == template_id:
            if not is_explicit_template(template):
                raise AudienceImportError(
                    f"Шаблон {template_id} не настроен как выборочный список"
                )
            return template
    raise AudienceImportError(f"Неизвестный шаблон: {template_id}")


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).replace("\n", " ").split()).strip().lower()
    return text.replace("ё", "е")


def _header_score(value: Any) -> int:
    normalized = _normalize_header(value)
    if not normalized:
        return 0

    best = 0
    for variant, score in EMAIL_HEADER_VARIANTS.items():
        if normalized == variant:
            best = max(best, score + 20)
        elif variant in normalized:
            best = max(best, score)
    return best


def _find_email_column(sheet, search_rows: int = 30) -> tuple[int, int, str]:
    candidates: list[tuple[int, int, int, str]] = []
    max_row = min(search_rows, sheet.max_row)

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            score = _header_score(value)
            if score:
                candidates.append((score, -row_idx, col_idx, str(value).strip()))

    if not candidates:
        raise AudienceImportError(
            "Не удалось найти колонку с email. Ожидается заголовок вида "
            "Email или Адрес электронной почты."
        )

    score, negative_row, col_idx, header = max(candidates)
    return -negative_row, col_idx, header


def extract_email_rows(file_bytes: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception as error:
        raise AudienceImportError(f"Не удалось открыть XLSX: {error}") from error

    sheet = workbook.active
    header_row, email_col, header_name = _find_email_column(sheet)
    rows: list[dict[str, Any]] = []

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        raw_value = sheet.cell(row=row_idx, column=email_col).value
        if raw_value is None or not str(raw_value).strip():
            continue

        source_email = str(raw_value).strip()
        normalized = normalize_email(source_email)
        rows.append(
            {
                "row_number": row_idx,
                "source_email": source_email,
                "normalized_email": normalized,
            }
        )

    workbook.close()

    if not rows:
        raise AudienceImportError("В найденной колонке нет ни одного заполненного email")

    metadata = {
        "sheet_name": sheet.title,
        "header_row": header_row,
        "email_column": email_col,
        "email_header": header_name,
    }
    return rows, metadata


def _safe_filename(filename: str) -> str:
    basename = Path(filename or "audience.xlsx").name
    cleaned = SAFE_FILENAME_RE.sub("_", basename).strip("._")
    if not cleaned:
        cleaned = "audience.xlsx"
    if not cleaned.lower().endswith(".xlsx"):
        cleaned += ".xlsx"
    return cleaned


def _row_result(
    connection,
    template_id: str,
    item: dict[str, Any],
    seen_emails: set[str],
) -> dict[str, Any]:
    normalized_email = item["normalized_email"]
    result = {
        **item,
        "status": "invalid",
        "error_text": "Некорректный формат email",
        "worker_key": None,
        "employment_seq": None,
        "employee_fio": None,
        "employee_email": None,
        "employee_login": None,
        "employee_department": None,
        "employee_position": None,
        "selected": 0,
    }

    if not normalized_email:
        return result

    if normalized_email in seen_emails:
        result["status"] = "duplicate"
        result["error_text"] = "Дубликат email в загруженном файле"
        return result
    seen_emails.add(normalized_email)

    active_matches = connection.execute(
        """
        SELECT * FROM employees
        WHERE active = 1 AND lower(trim(email)) = ?
        ORDER BY fio
        """,
        (normalized_email,),
    ).fetchall()

    if len(active_matches) > 1:
        result["status"] = "multiple_matches"
        result["error_text"] = "Email найден у нескольких активных сотрудников"
        return result

    if not active_matches:
        inactive_match = connection.execute(
            """
            SELECT * FROM employees
            WHERE active = 0 AND lower(trim(email)) = ?
            ORDER BY fio LIMIT 1
            """,
            (normalized_email,),
        ).fetchone()
        if inactive_match:
            result["status"] = "inactive_match"
            result["error_text"] = "Email найден только у неактивного сотрудника"
            result.update(
                {
                    "worker_key": inactive_match["worker_key"],
                    "employment_seq": inactive_match["employment_seq"],
                    "employee_fio": inactive_match["fio"],
                    "employee_email": inactive_match["email"],
                    "employee_login": inactive_match["login"],
                    "employee_department": inactive_match["department"],
                    "employee_position": inactive_match["position"],
                }
            )
        else:
            result["status"] = "not_found"
            result["error_text"] = "Email не найден среди сотрудников"
        return result

    employee = active_matches[0]
    result.update(
        {
            "worker_key": employee["worker_key"],
            "employment_seq": employee["employment_seq"],
            "employee_fio": employee["fio"],
            "employee_email": employee["email"],
            "employee_login": employee["login"],
            "employee_department": employee["department"],
            "employee_position": employee["position"],
        }
    )

    if not employee["login"]:
        result["status"] = "no_login"
        result["error_text"] = "У сотрудника отсутствует логин"
        return result

    assignment = connection.execute(
        """
        SELECT id FROM test_assignments
        WHERE template_id = ? AND worker_key = ? AND employment_seq = ? AND active = 1
        LIMIT 1
        """,
        (template_id, employee["worker_key"], employee["employment_seq"]),
    ).fetchone()

    if assignment:
        result["status"] = "already_assigned"
        result["error_text"] = "Тест уже назначен этому сотруднику"
        return result

    result["status"] = "ready"
    result["error_text"] = None
    result["selected"] = 1
    return result


def create_preview(
    settings: Settings,
    db: Database,
    template_id: str,
    original_filename: str,
    file_bytes: bytes,
    uploaded_by: str,
) -> dict[str, Any]:
    get_template(settings, template_id)

    if not original_filename.lower().endswith(".xlsx"):
        raise AudienceImportError("Поддерживаются только файлы XLSX")
    if not file_bytes:
        raise AudienceImportError("Загружен пустой файл")

    parsed_rows, metadata = extract_email_rows(file_bytes)
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target_dir = settings.data_path / "audiences" / template_id
    target_dir.mkdir(parents=True, exist_ok=True)
    saved_path = target_dir / f"{timestamp}_{_safe_filename(original_filename)}"
    saved_path.write_bytes(file_bytes)

    created_at = now_iso()
    with db.connect() as connection:
        cursor = connection.execute(
            """
            INSERT INTO audience_imports(
                template_id, original_filename, saved_path, file_hash, status,
                total_rows, created_at, uploaded_by, metadata_json
            ) VALUES (?, ?, ?, ?, 'preview', ?, ?, ?, ?)
            """,
            (
                template_id,
                original_filename,
                str(saved_path),
                file_hash,
                len(parsed_rows),
                created_at,
                uploaded_by,
                json.dumps(metadata, ensure_ascii=False),
            ),
        )
        import_id = int(cursor.lastrowid)

        seen_emails: set[str] = set()
        for item in parsed_rows:
            result = _row_result(connection, template_id, item, seen_emails)
            connection.execute(
                """
                INSERT INTO audience_import_rows(
                    import_id, row_number, source_email, normalized_email,
                    status, error_text, worker_key, employment_seq,
                    employee_fio, employee_email, employee_login,
                    employee_department, employee_position, selected
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    result["row_number"],
                    result["source_email"],
                    result["normalized_email"],
                    result["status"],
                    result["error_text"],
                    result["worker_key"],
                    result["employment_seq"],
                    result["employee_fio"],
                    result["employee_email"],
                    result["employee_login"],
                    result["employee_department"],
                    result["employee_position"],
                    result["selected"],
                ),
            )

        _refresh_import_counts(connection, import_id)

    return get_preview(db, import_id)


def _refresh_import_counts(connection, import_id: int) -> None:
    counts = Counter(
        row["status"]
        for row in connection.execute(
            "SELECT status FROM audience_import_rows WHERE import_id = ?", (import_id,)
        ).fetchall()
    )
    ready = counts.get("ready", 0)
    warning = counts.get("duplicate", 0) + counts.get("already_assigned", 0)
    errors = sum(
        counts.get(status, 0)
        for status in ("invalid", "not_found", "multiple_matches", "inactive_match", "no_login")
    )
    connection.execute(
        """
        UPDATE audience_imports
        SET ready_rows = ?, warning_rows = ?, error_rows = ?
        WHERE id = ?
        """,
        (ready, warning, errors, import_id),
    )


def _serialize_row(row) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


def get_preview(db: Database, import_id: int) -> dict[str, Any]:
    with db.connect() as connection:
        batch = connection.execute(
            "SELECT * FROM audience_imports WHERE id = ?", (import_id,)
        ).fetchone()
        if not batch:
            raise AudienceImportError("Импорт не найден")
        rows = connection.execute(
            """
            SELECT * FROM audience_import_rows
            WHERE import_id = ? ORDER BY row_number, id
            """,
            (import_id,),
        ).fetchall()

    return {
        "import": _serialize_row(batch),
        "rows": [_serialize_row(row) for row in rows],
    }


def confirm_preview(
    db: Database,
    import_id: int,
    selected_row_ids: Iterable[int],
) -> dict[str, Any]:
    selected_ids = {int(value) for value in selected_row_ids}
    confirmed_at = now_iso()
    imported = 0
    excluded = 0
    already_assigned = 0

    with db.connect() as connection:
        batch = connection.execute(
            "SELECT * FROM audience_imports WHERE id = ?", (import_id,)
        ).fetchone()
        if not batch:
            raise AudienceImportError("Импорт не найден")
        if batch["status"] != "preview":
            raise AudienceImportError("Этот импорт уже был подтвержден")

        ready_rows = connection.execute(
            """
            SELECT * FROM audience_import_rows
            WHERE import_id = ? AND status = 'ready'
            ORDER BY id
            """,
            (import_id,),
        ).fetchall()

        valid_ready_ids = {int(row["id"]) for row in ready_rows}
        unknown_ids = selected_ids - valid_ready_ids
        if unknown_ids:
            raise AudienceImportError("Выбраны строки, которые не готовы к импорту")

        for row in ready_rows:
            row_id = int(row["id"])
            if row_id not in selected_ids:
                connection.execute(
                    """
                    UPDATE audience_import_rows
                    SET status = 'excluded_by_operator', selected = 0,
                        error_text = 'Исключено оператором'
                    WHERE id = ?
                    """,
                    (row_id,),
                )
                excluded += 1
                continue

            cursor = connection.execute(
                """
                INSERT OR IGNORE INTO test_assignments(
                    template_id, worker_key, employment_seq,
                    source_import_id, source_row_id, assigned_at, active
                ) VALUES (?, ?, ?, ?, ?, ?, 1)
                """,
                (
                    batch["template_id"],
                    row["worker_key"],
                    row["employment_seq"],
                    import_id,
                    row_id,
                    confirmed_at,
                ),
            )

            if cursor.rowcount:
                connection.execute(
                    """
                    UPDATE audience_import_rows
                    SET status = 'imported', selected = 1, error_text = NULL
                    WHERE id = ?
                    """,
                    (row_id,),
                )
                imported += 1
            else:
                connection.execute(
                    """
                    UPDATE audience_import_rows
                    SET status = 'already_assigned', selected = 0,
                        error_text = 'Тест уже назначен этому сотруднику'
                    WHERE id = ?
                    """,
                    (row_id,),
                )
                already_assigned += 1

        connection.execute(
            """
            UPDATE audience_imports
            SET status = 'confirmed', confirmed_at = ?, confirmed_rows = ?
            WHERE id = ?
            """,
            (confirmed_at, imported, import_id),
        )
        _refresh_import_counts(connection, import_id)

    return {
        "import_id": import_id,
        "imported": imported,
        "excluded": excluded,
        "already_assigned": already_assigned,
        "preview": get_preview(db, import_id),
    }


def build_issues_workbook(db: Database, import_id: int) -> bytes:
    preview = get_preview(db, import_id)
    rows = [
        row
        for row in preview["rows"]
        if row["status"] not in {"ready", "imported"}
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ошибки импорта"
    sheet.append(
        [
            "Строка",
            "Email из файла",
            "Нормализованный email",
            "ФИО из основной базы",
            "Подразделение",
            "Должность",
            "Логин",
            "Статус",
            "Причина",
        ]
    )

    for row in rows:
        sheet.append(
            [
                row["row_number"],
                row["source_email"],
                row["normalized_email"],
                row["employee_fio"],
                row["employee_department"],
                row["employee_position"],
                row["employee_login"],
                STATUS_LABELS.get(row["status"], row["status"]),
                row["error_text"],
            ]
        )

    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 55)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
