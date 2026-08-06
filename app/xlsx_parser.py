from __future__ import annotations

import hashlib
import hmac
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SNILS_RE = re.compile(r"^\d{3}-?\d{3}-?\d{3}\s?\d{2}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
DEPARTMENT_SEPARATOR = " / "


@dataclass(frozen=True)
class EmployeePlacementRecord:
    """Одно кадровое назначение из строки XLSX."""

    department: str | None
    position: str | None


@dataclass
class EmployeeRecord:
    worker_key: str
    fio: str
    email: str | None
    login: str | None
    # Поля оставлены для обратной совместимости. Они содержат первое назначение.
    department: str | None
    position: str | None
    placements: tuple[EmployeePlacementRecord, ...] = ()


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip().lower()


def normalize_snils(value: Any) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    return digits if len(digits) == 11 else None


def worker_key(snils: str, secret: str) -> str:
    return hmac.new(secret.encode(), snils.encode(), hashlib.sha256).hexdigest()


def normalize_email(value: Any) -> str | None:
    if value is None:
        return None
    email = str(value).strip().lower()
    return email if EMAIL_RE.match(email) else None


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def find_header_row(sheet, expected: dict[str, str], search_rows: int) -> tuple[int, dict[str, int]]:
    normalized_expected = {key: normalize_header(name) for key, name in expected.items()}

    best_row = None
    best_mapping: dict[str, int] = {}
    for row_idx in range(1, min(search_rows, sheet.max_row) + 1):
        values = [
            normalize_header(sheet.cell(row=row_idx, column=col).value)
            for col in range(1, sheet.max_column + 1)
        ]
        mapping: dict[str, int] = {}
        for key, target in normalized_expected.items():
            for col_idx, value in enumerate(values, start=1):
                if value == target or target in value:
                    mapping[key] = col_idx
                    break
        if len(mapping) > len(best_mapping):
            best_row = row_idx
            best_mapping = mapping

    required = {"snils", "fio", "email"}
    if best_row is None or not required.issubset(best_mapping):
        missing = sorted(required - set(best_mapping))
        raise ValueError(f"Не найдены обязательные колонки XLSX: {', '.join(missing)}")

    return best_row, best_mapping


def _looks_like_department_row(values: list[Any], snils_col: int, fio_col: int, email_col: int) -> bool:
    snils_value = values[snils_col - 1] if snils_col <= len(values) else None
    fio_value = values[fio_col - 1] if fio_col <= len(values) else None
    email_value = values[email_col - 1] if email_col <= len(values) else None

    if normalize_snils(snils_value) or fio_value or email_value:
        return False

    nonempty = [_normalize_text(value) for value in values if _normalize_text(value)]
    return len(nonempty) == 1


def _department_path(hierarchy: dict[int, str]) -> str | None:
    """Формирует полный путь и убирает только соседние одинаковые уровни."""
    parts: list[str] = []

    for level in sorted(hierarchy):
        department = _normalize_text(hierarchy[level])
        if not department:
            continue

        if parts and parts[-1].casefold() == department.casefold():
            continue

        parts.append(department)

    return DEPARTMENT_SEPARATOR.join(parts) or None


def _update_department_hierarchy(
    hierarchy: dict[int, str],
    outline_level: int,
    department: str,
) -> str | None:
    """Обновляет стек подразделений при переходе на новый уровень XLSX."""
    for level in list(hierarchy):
        if level >= outline_level:
            del hierarchy[level]

    hierarchy[outline_level] = department
    return _department_path(hierarchy)


def _placement_key(placement: EmployeePlacementRecord) -> tuple[str, str]:
    return (
        _normalize_text(placement.department).casefold(),
        _normalize_text(placement.position).casefold(),
    )


def _append_placement(
    placements: tuple[EmployeePlacementRecord, ...],
    placement: EmployeePlacementRecord,
) -> tuple[EmployeePlacementRecord, ...]:
    """Добавляет назначение, удаляя только полностью одинаковую пару."""
    key = _placement_key(placement)
    if key in {_placement_key(item) for item in placements}:
        return placements
    return (*placements, placement)


def parse_xlsx(
    path: Path,
    columns: dict[str, str],
    header_search_rows: int,
    hash_secret: str,
    login_overrides: dict[str, str],
    sheet_name: str | None = None,
    lowercase_login: bool = True,
) -> list[EmployeeRecord]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    header_row, mapping = find_header_row(sheet, columns, header_search_rows)
    snils_col = mapping["snils"]
    fio_col = mapping["fio"]
    email_col = mapping["email"]
    position_col = mapping.get("position")

    department_hierarchy: dict[int, str] = {}
    current_department: str | None = None
    merged: dict[str, EmployeeRecord] = {}

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        values = [
            sheet.cell(row=row_idx, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

        if _looks_like_department_row(values, snils_col, fio_col, email_col):
            department = next(
                _normalize_text(value)
                for value in values
                if _normalize_text(value)
            )
            outline_level = int(sheet.row_dimensions[row_idx].outlineLevel or 0)
            current_department = _update_department_hierarchy(
                department_hierarchy,
                outline_level,
                department,
            )
            continue

        snils = normalize_snils(values[snils_col - 1])
        fio_raw = values[fio_col - 1]
        if not snils or fio_raw is None or not str(fio_raw).strip():
            continue

        fio = _normalize_text(fio_raw)
        email = normalize_email(values[email_col - 1])
        login = None
        if email:
            login = login_overrides.get(email) or email.split("@", 1)[0]
            login = login.strip()
            if lowercase_login:
                login = login.lower()

        position = None
        if position_col:
            raw_position = values[position_col - 1]
            if raw_position is not None and str(raw_position).strip():
                position = _normalize_text(raw_position)

        placement = EmployeePlacementRecord(
            department=current_department,
            position=position,
        )
        key = worker_key(snils, hash_secret)
        existing = merged.get(key)

        if existing:
            placements = _append_placement(existing.placements, placement)
            primary = placements[0] if placements else placement
            merged[key] = EmployeeRecord(
                worker_key=key,
                fio=fio or existing.fio,
                email=email or existing.email,
                login=login or existing.login,
                department=primary.department,
                position=primary.position,
                placements=placements,
            )
        else:
            merged[key] = EmployeeRecord(
                worker_key=key,
                fio=fio,
                email=email,
                login=login,
                department=placement.department,
                position=placement.position,
                placements=(placement,),
            )

    workbook.close()
    return list(merged.values())
