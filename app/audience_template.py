from __future__ import annotations

import io
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


TEMPLATE_FILENAME = "Шаблон_списка_участников.xlsx"

TEMPLATE_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def build_audience_template() -> bytes:
    workbook = Workbook()

    participants = workbook.active
    participants.title = "Участники"
    participants.sheet_view.showGridLines = False
    participants.freeze_panes = "A2"

    headers = (
        "ФИО",
        "Должность",
        "Email",
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="DCE6F1",
    )

    header_font = Font(
        name="Arial",
        size=11,
        bold=True,
        color="1F2937",
    )

    thin_side = Side(
        style="thin",
        color="BFC7D1",
    )

    header_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        cell = participants.cell(
            row=1,
            column=column_number,
            value=header,
        )

        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    participants.row_dimensions[1].height = 26

    participants.column_dimensions["A"].width = 42
    participants.column_dimensions["B"].width = 42
    participants.column_dimensions["C"].width = 38

    participants.auto_filter.ref = "A1:C10000"

    participants["A1"].comment = Comment(
        "ФИО указывается для удобства проверки списка. "
        "При импорте система берет актуальное ФИО из основной базы.",
        "Система тестирования",
    )

    participants["B1"].comment = Comment(
        "Должность указывается для удобства проверки списка. "
        "При импорте система берет актуальную должность из основной базы.",
        "Система тестирования",
    )

    participants["C1"].comment = Comment(
        "Обязательное поле. Укажите один адрес электронной почты "
        "работника в каждой строке.",
        "Система тестирования",
    )

    email_validation = DataValidation(
        type="custom",
        formula1=(
            '=OR(C2="",'
            'AND('
            'ISNUMBER(SEARCH("@",C2)),'
            'ISNUMBER(SEARCH(".",C2,SEARCH("@",C2)+2))'
            ')'
            ')'
        ),
        allow_blank=True,
    )

    email_validation.errorTitle = "Некорректный Email"
    email_validation.error = (
        "Укажите один корректный адрес электронной почты."
    )

    email_validation.promptTitle = "Email работника"
    email_validation.prompt = (
        "Укажите один адрес электронной почты в этой строке."
    )

    email_validation.showErrorMessage = True
    email_validation.showInputMessage = True

    participants.add_data_validation(email_validation)
    email_validation.add("C2:C10000")

    instructions = workbook.create_sheet("Инструкция")
    instructions.sheet_view.showGridLines = False
    instructions.column_dimensions["A"].width = 115

    instructions["A1"] = "Шаблон списка участников тестирования"
    instructions["A1"].font = Font(
        name="Arial",
        size=16,
        bold=True,
        color="1F2937",
    )

    instructions["A3"] = (
        "1. Заполняйте только лист «Участники»."
    )

    instructions["A4"] = (
        "2. В каждой строке должен быть указан один работник."
    )

    instructions["A5"] = (
        "3. Поле Email является обязательным."
    )

    instructions["A6"] = (
        "4. ФИО и должность указываются для удобства проверки "
        "и координации списка."
    )

    instructions["A7"] = (
        "5. При импорте система использует Email для поиска работника."
    )

    instructions["A8"] = (
        "6. Актуальные ФИО, должность, подразделение и логин "
        "берутся из основной базы сотрудников."
    )

    instructions["A9"] = (
        "7. Не изменяйте название листа «Участники» "
        "и заголовки столбцов."
    )

    instructions["A10"] = (
        "8. Не указывайте несколько адресов электронной почты "
        "в одной ячейке."
    )

    instructions["A11"] = (
        "9. Не объединяйте ячейки и не добавляйте дополнительные "
        "строки над заголовками."
    )

    instructions["A12"] = (
        "10. Пустые строки и повторяющиеся адреса будут проверены "
        "системой при загрузке."
    )

    for row_number in range(3, 13):
        cell = instructions.cell(
            row=row_number,
            column=1,
        )

        cell.font = Font(
            name="Arial",
            size=11,
            color="344054",
        )

        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        instructions.row_dimensions[row_number].height = 25

    workbook.active = 0

    output = io.BytesIO()
    workbook.save(output)

    return output.getvalue()


def template_download_headers() -> dict[str, str]:
    encoded_filename = quote(
        TEMPLATE_FILENAME,
        safe="",
    )

    return {
        "Content-Disposition": (
            'attachment; filename="audience_template.xlsx"; '
            f"filename*=UTF-8''{encoded_filename}"
        ),
        "Cache-Control": "no-store",
    }